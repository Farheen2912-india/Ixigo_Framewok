import openpyxl

def get_data():
    workbook = openpyxl.load_workbook("C://Users//HP//PycharmProjects//selenium_E26//Ixigo_Login_Framework//test_data//login_ixigodata.xlsx")

    worksheet = workbook["Sheet1"]
    data = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        print("ROW:", row)
        if row[0] is None and row[1] is None:
            continue
        data.append(row)
    print("DATA:", data)
    return data

def get_flight_data():

    workbook = openpyxl.load_workbook(
        "C://Users//HP//PycharmProjects//selenium_E26//Ixigo_Login_Framework//test_data//flight_data.xlsx"
    )

    worksheet = workbook["FlightSearchData"]

    data = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):

        print("FLIGHT ROW:", row)

        if row[0] is None and row[1] is None:
            continue

        data.append(row)

    print("FLIGHT DATA:", data)

    return data